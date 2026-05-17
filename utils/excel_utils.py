from openpyxl import load_workbook


def get_login_test_data():

    workbook = load_workbook(
        "testdata/login_test_data.xlsx"
    )

    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        data.append(row)

    return data
